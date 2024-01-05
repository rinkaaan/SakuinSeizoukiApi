import os
import re

import fitz
import openpyxl
from apiflask import APIBlueprint, Schema
from marshmallow.fields import String, List, Dict, Nested, Integer
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

from api.resources.word_list import parse_word_list

from api.schemas.main import PageTypeDetail, WordPages
from nguylinc_python_utils.pyinstaller import get_bundle_dir

index_bp = APIBlueprint("Index", __name__, url_prefix="/index")


class CreateIndexIn(Schema):
    list_path = String()
    pdf_path = String()
    sheet_name = String()
    start_cell = String()
    end_cell = String()
    page_types = Dict(String(), Nested(PageTypeDetail))


class CreateIndexOut(Schema):
    word_pages = List(Nested(WordPages))
    missing_pages = List(Integer())
    missing_words = List(String())


@index_bp.post("/create")
@index_bp.input(CreateIndexIn, arg_name="params")
@index_bp.output(CreateIndexOut)
def create_index(params):
    words = parse_word_list(params["list_path"], params["sheet_name"], params["start_cell"], params["end_cell"])
    # Dictionary to store page number and content
    page_contents = {}
    # Dictionary to hold words and the pages they appear on
    word_pages = {}
    # List to store pages that were not found
    missing_pages = []
    # List to store words not found on any pages
    missing_words = []
    # Dictionary to hold filtered word_pages where all words have at least one page
    filtered_word_pages = {}

    doc = fitz.open(params["pdf_path"])

    # Process each page type
    for page_type, details in params["page_types"].items():
        page_numbers = details["page_numbers"]
        for page_number in page_numbers:
            page = doc.load_page(page_number - 1)
            current_group_index = -1
            content = ""

            for annotation in details["annotations"]:
                x, y = annotation["x"], annotation["y"]
                width, height = annotation["width"], annotation["height"]
                rect = fitz.Rect(x, y, x + width, y + height)
                chars = page.get_textbox(rect)

                if annotation["group_index"] == current_group_index:
                    is_first_annotation = False
                else:
                    is_first_annotation = True
                    current_group_index = annotation["group_index"]
                    finished_group = False

                if finished_group:
                    continue
                if is_first_annotation:
                    content = "".join(char for char in chars if ord(char) > 32)
                else:
                    actual_page_number = re.sub(r"[^0-9]", "", chars)
                    if actual_page_number and content:
                        page_contents[int(actual_page_number)] = content
                        finished_group = True

    # For each word, iterate through the pages and check if the word is in the page content and update the word_pages dictionary
    # For each page, also combine contents of next page to see if word is being split across pages
    for word in words:
        word_pages[word] = []
        for page_number, content in page_contents.items():
            # For each word, iterate through the pages and check if the word is in the page content and update the word_pages dictionary
            # For each page, also combine contents of next page to see if word is being split across pages

            current_page_len = len(content)
            if page_number + 1 in page_contents:
                content += page_contents[page_number + 1]
            if word in content and content.index(word) < current_page_len:
                word_pages[word].append(page_number)

    # Check if any pages are missing; set range from 1 to max of all keys in page_contents
    for page_number in range(1, max(page_contents.keys()) + 1):
        if page_number not in page_contents:
            missing_pages.append(page_number)

    # Remove empty arrays from word_pages and update missing_words
    for word, pages in word_pages.items():
        if not pages:
            missing_words.append(word)
        else:
            filtered_word_pages[word] = pages

    # Reshape word_pages to list of objects
    filtered_word_pages = [{"word": word, "pages": pages} for word, pages in filtered_word_pages.items()]

    return {
        "word_pages": filtered_word_pages,
        "missing_pages": missing_pages,
        "missing_words": missing_words,
    }


class GetIndexOut(Schema):
    url = String()


@index_bp.post("/get")
@index_bp.input(CreateIndexOut, arg_name="params")
@index_bp.output(GetIndexOut)
def get_index(params):
    workbook = Workbook()
    sheet1 = workbook.active
    # sheet1.title = "Index"
    sheet1.title = "索引"

    # sheet1["A1"] = "Word"
    # sheet1["B1"] = "Pages"
    sheet1["A1"] = "索引語"
    sheet1["B1"] = "ページ"
    for cell in sheet1["1:1"]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Add words with pages
    for index, word_pages in enumerate(params["word_pages"]):
        sheet1[f"A{index + 2}"] = word_pages["word"]
        sheet1[f"B{index + 2}"] = ", ".join(str(page) for page in word_pages["pages"])

    # Resize column A to fit the longest word
    dim_holder_1 = DimensionHolder(worksheet=sheet1)
    longest_word = ""
    for word_pages in params["word_pages"]:
        if len(word_pages["word"]) > len(longest_word):
            longest_word = word_pages["word"]
    if len("索引語") > len(longest_word):
        longest_word = "索引語"
    dim_holder_1["A"] = ColumnDimension(sheet1, min=1, max=1, width=len(longest_word) * 2.1)
    sheet1.column_dimensions = dim_holder_1

    # sheet2 = workbook.create_sheet("Missing Words")
    sheet2 = workbook.create_sheet("見つからない言葉")
    # sheet2["A1"] = "Missing words"
    sheet2["A1"] = "見つからない言葉"
    for cell in sheet2["1:1"]:
        cell.font = openpyxl.styles.Font(bold=True)
    # Add missing words below header
    for index, word in enumerate(params["missing_words"]):
        sheet2[f"A{index + 2}"] = word

    # Resize column A to fit the longest word
    dim_holder_2 = DimensionHolder(worksheet=sheet2)
    longest_word = ""
    for word in params["missing_words"]:
        if len(word) > len(longest_word):
            longest_word = word
    if len("見つからない言葉") > len(longest_word):
        longest_word = "見つからない言葉"
    dim_holder_2["A"] = ColumnDimension(sheet2, min=1, max=1, width=len(longest_word) * 2.1)
    sheet2.column_dimensions = dim_holder_2

    # sheet3 = workbook.create_sheet("Missing Pages")
    sheet3 = workbook.create_sheet("見つからないページ")
    # sheet3["A1"] = "Missing pages"
    sheet3["A1"] = "見つからないページ"
    for cell in sheet3["1:1"]:
        cell.font = openpyxl.styles.Font(bold=True)
    # Add missing pages
    for index, page in enumerate(params["missing_pages"]):
        sheet3[f"A{index + 2}"] = page

    # Resize column A to fit the longest page number
    dim_holder_3 = DimensionHolder(worksheet=sheet3)
    longest_page_number = ""
    for page in params["missing_pages"]:
        if len(str(page)) > len(longest_page_number):
            longest_page_number = str(page)
    if (len("見つからないページ") > len(longest_page_number)):
        longest_page_number = "見つからないページ"
    dim_holder_3["A"] = ColumnDimension(sheet3, min=1, max=1, width=len(longest_page_number) * 2.1)
    sheet3.column_dimensions = dim_holder_3

    # Remove temporary file if it exists
    try:
        os.remove(get_bundle_dir() + "/temp/index.xlsx")
    except FileNotFoundError:
        pass

    # Save the workbook to a temporary file
    # workbook.save(get_bundle_dir() + "/temp/index.xlsx")
    workbook.save(get_bundle_dir() + "/temp/索引.xlsx")
    return {
        "url": "/temp/索引.xlsx"
    }
