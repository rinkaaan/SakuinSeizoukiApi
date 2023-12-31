import io
import os
import re

import fitz
import openpyxl
from apiflask import APIBlueprint, Schema
from apiflask.fields import String, Integer, List, Nested, Dict
from flask import send_file
from openpyxl.workbook import Workbook

from nguylinc_python_utils.pyinstaller import get_bundle_dir

project_bp = APIBlueprint("Project", __name__, url_prefix="/project")


class OpenPdfIn(Schema):
    pdf_path = String()


class PdfPageType(Schema):
    width = Integer()
    height = Integer()
    page_numbers = List(Integer())
    type = Integer()


class OpenPdfOut(Schema):
    page_types = List(Nested(PdfPageType))


@project_bp.post("/new/pdf")
@project_bp.input(OpenPdfIn, arg_name="params")
@project_bp.output(OpenPdfOut)
def open_pdf(params):
    doc = fitz.open(params["pdf_path"])

    # Initialize an empty dictionary to store page types and their sample page numbers
    page_types = {}

    # Loop through all pages in the document
    for page_number, page in enumerate(doc):
        # Get the current page's width and height; rounded up
        page_width = int(round(page.mediabox_size[0]))
        page_height = int(round(page.mediabox_size[1]))

        # Check if the current page size exists in the `page_types` dictionary
        if (page_width, page_height) not in page_types:
            # Create a new entry for the current page size
            page_types[(page_width, page_height)] = {
                "width": page_width,
                "height": page_height,
                "page_numbers": [],
            }

        # Append the current page number to the sample page numbers for the current page size
        page_types[(page_width, page_height)]["page_numbers"].append(page_number + 1)

    # Convert the dictionary of page types to a list of page types
    page_types = list(page_types.values())

    # Sort the list of page types by the number of sample page numbers in descending order
    page_types.sort(key=lambda x: len(x["page_numbers"]), reverse=True)

    # Assign a type number to each page type; starting from 0
    for index, page_type in enumerate(page_types):
        page_type["type"] = index

    return {"page_types": page_types}


class GetPdfPageIn(Schema):
    pdf_path = String()
    page_number = Integer()


@project_bp.get("/get/pdf/page")
@project_bp.input(GetPdfPageIn, arg_name="params", location="query")
@project_bp.output({}, content_type="image/png")
def get_pdf_page(params):
    doc = fitz.open(params["pdf_path"])
    page = doc.load_page(params["page_number"] - 1)
    zoom = 1
    mat = fitz.Matrix(zoom, zoom)
    image = page.get_pixmap(matrix=mat).tobytes()
    return send_file(io.BytesIO(image), mimetype="image/png")


class Annotation(Schema):
    x = Integer()
    y = Integer()
    width = Integer()
    height = Integer()
    group_index = Integer()


class PageTypeDetail(Schema):
    annotations = List(Nested(Annotation))
    page_numbers = List(Integer())


class CreateIndexIn(Schema):
    list_path = String()
    pdf_path = String()
    sheet_name = String()
    start_cell = String()
    end_cell = String()
    page_types = Dict(String(), Nested(PageTypeDetail))


class WordPages(Schema):
    word = String()
    pages = List(Integer())


class CreateIndexOut(Schema):
    word_pages = List(Nested(WordPages))
    missing_pages = List(Integer())
    missing_words = List(String())


@project_bp.post("/create/index")
@project_bp.input(CreateIndexIn, arg_name="params")
@project_bp.output(CreateIndexOut)
def create_index(params):
    words = parse_word_list(params["list_path"], params["sheet_name"], params["start_cell"], params["end_cell"])
    # Dictionary to store page number and content
    page_contents = {}
    # Dictionary to hold words and the pages they appear on
    word_pages = {}
    # List to store pages that were not found
    missing_pages = []
    # List to store words not found on any pages
    missing_words = []
    # Dictionary to hold filtered word_pages where all words have at least one page
    filtered_word_pages = {}

    doc = fitz.open(params["pdf_path"])

    # Process each page type
    for page_type, details in params["page_types"].items():
        page_numbers = details["page_numbers"]
        for page_number in page_numbers:
            page = doc.load_page(page_number - 1)
            current_group_index = -1
            content = ""

            for annotation in details["annotations"]:
                x, y = annotation["x"], annotation["y"]
                width, height = annotation["width"], annotation["height"]
                rect = fitz.Rect(x, y, x + width, y + height)
                chars = page.get_textbox(rect)

                if annotation["group_index"] == current_group_index:
                    is_first_annotation = False
                else:
                    is_first_annotation = True
                    current_group_index = annotation["group_index"]
                    finished_group = False

                if finished_group:
                    continue
                if is_first_annotation:
                    content = "".join(char for char in chars if ord(char) > 32)
                else:
                    actual_page_number = re.sub(r"[^0-9]", "", chars)
                    if actual_page_number and content:
                        page_contents[int(actual_page_number)] = content
                        finished_group = True

    # For each word, iterate through the pages and check if the word is in the page content and update the word_pages dictionary
    # For each page, also combine contents of next page to see if word is being split across pages
    for word in words:
        word_pages[word] = []
        for page_number, content in page_contents.items():
            # For each word, iterate through the pages and check if the word is in the page content and update the word_pages dictionary
            # For each page, also combine contents of next page to see if word is being split across pages

            current_page_len = len(content)
            if page_number + 1 in page_contents:
                content += page_contents[page_number + 1]
            if word in content and content.index(word) < current_page_len:
                word_pages[word].append(page_number)

    # Check if any pages are missing; set range from 1 to max of all keys in page_contents
    for page_number in range(1, max(page_contents.keys()) + 1):
        if page_number not in page_contents:
            missing_pages.append(page_number)

    # Remove empty arrays from word_pages and update missing_words
    for word, pages in word_pages.items():
        if not pages:
            missing_words.append(word)
        else:
            filtered_word_pages[word] = pages

    # Reshape word_pages to list of objects
    filtered_word_pages = [{"word": word, "pages": pages} for word, pages in filtered_word_pages.items()]

    return {
        "word_pages": filtered_word_pages,
        "missing_pages": missing_pages,
        "missing_words": missing_words,
    }


class GetWordListIn(Schema):
    word_list_path = String()
    sheet_name = String()
    start_cell = String()
    end_cell = String()


class GetWordListOut(Schema):
    word_list = List(String())


@project_bp.get("/get/word/list")
@project_bp.input(GetWordListIn, arg_name="params", location="query")
@project_bp.output(GetWordListOut)
def get_word_list(params):
    word_list_path = params["word_list_path"]
    sheet_name = params["sheet_name"]
    start_cell = params["start_cell"]
    end_cell = params["end_cell"]

    return {"word_list": parse_word_list(word_list_path, sheet_name, start_cell, end_cell)}


def parse_word_list(word_list_path, sheet_name, start_cell, end_cell):
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(word_list_path)
    sheet = workbook[sheet_name]

    # Initialize an empty list to store cell values
    cell_values = []

    # Define the end row for reading
    end_row = sheet.max_row if end_cell is None else sheet[end_cell].row

    # Iterate through the cells in the specified column from start_cell to end_cell
    for row in range(sheet[start_cell].row, end_row + 1):
        cell = sheet[f"{start_cell[0]}{row}"]
        if (cell.value is None) or (cell.value == ""):
            continue
        cell_values.append(str(cell.value))

    return cell_values


class GetIndexOut(Schema):
    url = String()


@project_bp.post("/get/index")
@project_bp.input(CreateIndexOut, arg_name="params")
@project_bp.output(GetIndexOut)
def get_index(params):
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Word"
    sheet["B1"] = "Pages"

    for index, word_pages in enumerate(params["word_pages"]):
        sheet[f"A{index + 2}"] = word_pages["word"]
        sheet[f"B{index + 2}"] = ", ".join(str(page) for page in word_pages["pages"])

    # Adding words with no pages at end
    for index, word in enumerate(params["missing_words"]):
        sheet[f"A{index + 2 + len(params['word_pages'])}"] = word
        sheet[f"B{index + 2 + len(params['word_pages'])}"] = "No pages found"

    # Remove temporary file if it exists
    try:
        os.remove(get_bundle_dir() + "/temp/index.xlsx")
    except FileNotFoundError:
        pass

    # Save the workbook to a temporary file
    workbook.save(get_bundle_dir() + "/temp/index.xlsx")
    return {
        "url": "/temp/index.xlsx"
    }
